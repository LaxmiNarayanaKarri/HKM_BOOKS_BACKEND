"""
xlsx streaming helpers.

Replaces the Flask-era `app_shared.xlxs.xlsx_export` (`send_xlsx`,
`send_xlsx_multi`), which returned Flask `send_file` responses built
around the request/response cycle of a template-rendering app. These
build the same workbook shapes but hand back a FastAPI
`StreamingResponse`, so the routers stay pure-JSON-API elsewhere and
only these two export helpers deal with binary output.
"""

from dataclasses import is_dataclass
from io import BytesIO
from typing import Any, Iterable, List, Optional, Sequence, Tuple

from fastapi.responses import StreamingResponse
from openpyxl import Workbook

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _row_value(row: Any, column: str) -> Any:
    if is_dataclass(row):
        return getattr(row, column, "")
    if isinstance(row, dict):
        return row.get(column, "")
    return getattr(row, column, "")


def _write_sheet(ws, rows: Iterable[Any], columns: Sequence[str], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([_row_value(row, col) for col in columns])


def _attachment_headers(filename: str) -> dict:
    return {"Content-Disposition": f'attachment; filename="{filename}"'}


def send_xlsx(rows: Iterable[Any], columns: Sequence[str], headers: Sequence[str],
              sheet_name: str, filename: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    _write_sheet(ws, rows, columns, headers)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type=XLSX_MEDIA_TYPE, headers=_attachment_headers(filename))


def send_xlsx_multi(sheets: List[Tuple[str, Iterable[Any], Sequence[str], Optional[Sequence[str]]]],
                     filename: str) -> StreamingResponse:
    """`sheets` is a list of (sheet_name, rows, columns, headers) tuples.
    `headers` defaults to a title-cased version of `columns` when None,
    matching the original helper's behaviour."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows, columns, headers in sheets:
        ws = wb.create_sheet(title=sheet_name[:31] or "Sheet")
        _write_sheet(ws, rows, columns, headers or [c.replace("_", " ").title() for c in columns])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type=XLSX_MEDIA_TYPE, headers=_attachment_headers(filename))
